"""
Excel QA Extractor & Fine-tuning Formatter
--------------------------------------------
Parses the NUST Bank product knowledge Excel workbook, detects question /
answer rows using heuristics, and writes three output files:

  1. data/all_qa_pairs.json         -- every extracted QA pair (JSON array)
  2. data/finetuning_data.jsonl     -- Alpaca-style instruction format
  3. data/finetuning_data_chat.jsonl -- OpenAI chat-completion format

The workbook contains 34+ sheets with varying layouts, merged cells, and
multi-row answers.  The script handles all of these edge-cases.
"""

import json
import os
import re
from collections import Counter

import openpyxl

# ── Path resolution ──────────────────────────────────────────────────────
_HERE        = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_HERE)

WORKBOOK_PATH     = os.path.join(_PROJECT_ROOT, "NUST Bank-Product-Knowledge.xlsx")
OUT_ALL_QA_JSON   = os.path.join(_HERE, "all_qa_pairs.json")
OUT_INSTRUCT_JSONL = os.path.join(_HERE, "finetuning_data.jsonl")
OUT_CHAT_JSONL    = os.path.join(_HERE, "finetuning_data_chat.jsonl")

IGNORED_SHEETS = {"Main", "Rate Sheet July 1 2024", "Sheet1"}

ASSISTANT_PERSONA = (
    "You are a helpful customer support assistant for NUST Bank. "
    "Answer the customer's question accurately and concisely using "
    "the bank's product knowledge."
)


# ── Text utilities ───────────────────────────────────────────────────────

def tidy(text: str) -> str:
    """Normalise whitespace and remove tabs while preserving paragraph breaks."""
    if not text:
        return ""
    text = text.strip().replace("\t", " ")
    text = re.sub(r"[^\S\n]+", " ", text)      # collapse horizontal ws
    text = re.sub(r"\n{3,}", "\n\n", text)      # limit consecutive blank lines
    return text.strip()


def looks_like_question(text: str) -> bool:
    """Heuristic to decide whether *text* is a question row."""
    s = text.strip()
    if not s:
        return False

    if s.endswith("?"):
        return True

    QUESTION_PREFIXES = (
        "what", "how", "is ", "is\n", "can ", "can\n", "does", "do ",
        "are ", "which", "who", "where", "when", "why",
        "i would like to", "i want to", "please tell",
        "1.", "1 .", "1-",
    )
    low = s.lower()
    if any(low.startswith(pfx) for pfx in QUESTION_PREFIXES):
        return True

    # question mark within the first 80 chars (merged-cell artefact)
    if "?" in s[:80]:
        return True

    return False


# ── Sheet-level extraction ───────────────────────────────────────────────

def _read_product_title(ws) -> str:
    """Derive a product name from the first row of the worksheet."""
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            candidate = str(cell.value).strip()
            if candidate.startswith("=") or candidate.lower() == "main":
                continue
            if candidate:
                return candidate
    return ""


def _collect_row_text(ws) -> dict[int, str]:
    """Return {row_number: combined_text} for every non-empty row."""
    mapping: dict[int, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        parts = []
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if not val or val.startswith("=") or val.lower() == "main":
                continue
            parts.append(val)
        if parts:
            joined = " | ".join(parts) if len(parts) > 1 else parts[0]
            row_num = row.row if hasattr(row, "row") else row[0].row
            mapping[row_num] = joined
    return mapping


def extract_pairs(ws, sheet_label: str) -> list[dict]:
    """Walk through *ws* detecting question/answer boundaries."""
    product = _read_product_title(ws) or sheet_label

    row_texts = _collect_row_text(ws)
    if not row_texts:
        return []

    ordered = sorted(row_texts.items())

    # skip first row when it is just the product title
    offset = 1 if ordered and ordered[0][1].strip() == product.strip() else 0

    pairs: list[dict] = []
    pending_q: str | None = None
    pending_a_lines: list[str] = []

    def _flush():
        nonlocal pending_q, pending_a_lines
        if pending_q and pending_a_lines:
            combined = tidy("\n".join(pending_a_lines))
            if combined:
                pairs.append({
                    "question": tidy(pending_q),
                    "answer":   combined,
                    "product":  product,
                    "sheet":    sheet_label,
                })
        pending_q = None
        pending_a_lines = []

    for _, text in ordered[offset:]:
        text = tidy(text)
        if not text:
            continue

        if looks_like_question(text):
            _flush()
            pending_q = text
        elif pending_q is not None:
            pending_a_lines.append(text)

    _flush()  # handle the final pair
    return pairs


# ── JSON FAQ supplement ──────────────────────────────────────────────────

def _ingest_json_faq(path: str) -> list[dict]:
    """Load the supplementary JSON FAQ file and normalise entries."""
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as fh:
        blob = json.load(fh)

    extra: list[dict] = []
    for cat in blob.get("categories", []):
        label = cat.get("category", "General")
        for item in cat.get("questions", []):
            extra.append({
                "question": tidy(item["question"]),
                "answer":   tidy(item["answer"]),
                "product":  f"Mobile App - {label}",
                "sheet":    "funds_transfer_app_features_faq.json",
            })
    return extra


# ── Output writers ───────────────────────────────────────────────────────

def _write_all_qa_json(qa_list: list[dict], dest: str):
    with open(dest, "w", encoding="utf-8") as fh:
        json.dump(qa_list, fh, indent=2, ensure_ascii=False)
    print(f"  -> {dest}  ({len(qa_list)} pairs)")


def _write_instruct_jsonl(qa_list: list[dict], dest: str):
    with open(dest, "w", encoding="utf-8") as fh:
        for qa in qa_list:
            fh.write(json.dumps({
                "instruction": qa["question"],
                "input":       "",
                "output":      qa["answer"],
                "product":     qa["product"],
            }, ensure_ascii=False) + "\n")
    print(f"  -> {dest}")


def _write_chat_jsonl(qa_list: list[dict], dest: str):
    with open(dest, "w", encoding="utf-8") as fh:
        for qa in qa_list:
            fh.write(json.dumps({
                "messages": [
                    {"role": "system",    "content": ASSISTANT_PERSONA},
                    {"role": "user",      "content": qa["question"]},
                    {"role": "assistant", "content": qa["answer"]},
                ],
            }, ensure_ascii=False) + "\n")
    print(f"  -> {dest}")


# ── Report ───────────────────────────────────────────────────────────────

def _print_report(qa_list: list[dict]):
    freq = Counter(qa["product"] for qa in qa_list)
    bar = "-" * 58

    print(f"\n{bar}")
    print(f"{'Product':<44} {'Count':>10}")
    print(bar)
    for name, cnt in freq.most_common():
        print(f"  {name:<42} {cnt:>10}")
    print(bar)
    print(f"  {'TOTAL':<42} {len(qa_list):>10}")

    print(f"\n{'=' * 58}")
    print("Sample entries (first 3):")
    print("=" * 58)
    for i, qa in enumerate(qa_list[:3], start=1):
        print(f"\n  [{i}] Product: {qa['product']}")
        print(f"      Q: {qa['question'][:120]}")
        print(f"      A: {qa['answer'][:200]}")


# ── Entry point ──────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

    all_pairs: list[dict] = []
    skipped: list[str] = []

    for name in wb.sheetnames:
        if name in IGNORED_SHEETS:
            skipped.append(name)
            continue
        sheet_pairs = extract_pairs(wb[name], name)
        all_pairs.extend(sheet_pairs)
        print(f"  + {name:25s}  {len(sheet_pairs):>3} pairs")

    print(f"\nIgnored sheets: {skipped}")
    print(f"Extracted from workbook: {len(all_pairs)} pairs")

    # merge supplementary JSON FAQ
    faq_path = os.path.join(_PROJECT_ROOT, "funds_transer_app_features_faq.json")
    supplementary = _ingest_json_faq(faq_path)
    if supplementary:
        all_pairs.extend(supplementary)
        print(f"  + {'JSON FAQ':25s}  {len(supplementary):>3} pairs")
        print(f"Combined total: {len(all_pairs)} pairs")

    # persist
    _write_all_qa_json(all_pairs, OUT_ALL_QA_JSON)
    _write_instruct_jsonl(all_pairs, OUT_INSTRUCT_JSONL)
    _write_chat_jsonl(all_pairs, OUT_CHAT_JSONL)

    _print_report(all_pairs)


if __name__ == "__main__":
    main()
